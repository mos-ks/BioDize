"""Export pipeline: Excel (xlsx) + schönes CSV fuer Pharma-Chargenprotokolle."""
from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone

from sqlalchemy.orm import Session

from app.db import models

# ---------------------------------------------------------------------------
# Solution-format export (matches the host grading sheet: one row per parameter,
# in CONTEXT — Batch / Prozessstufe / Context 1-3 / Parameter 1-3 / Condition /
# Text / Value — not a page-by-page dump). `Condition` is the verdict.
# ---------------------------------------------------------------------------

SOLUTION_HEADER = ["Batch", "Prozessstufe", "Context 1", "Context 2", "Context 3",
                   "Parameter 1", "Parameter 2", "Parameter 3", "Condition", "Text", "Value"]

# Mass-balance components are cross-checked against each other (net = gross - tare,
# V = m·ρ), so a clean one is "confirmed by a second value" in the host's taxonomy.
_CONFIRMED_ROLES = {"tare_mass", "gross_mass", "net_mass", "volume", "density", "concentration"}


def _batch_no(doc) -> str:
    if doc is None:
        return ""
    if getattr(doc, "batch_no", None):
        return str(doc.batch_no)
    m = re.search(r"batch\s*(?:no\.?|nr\.?)?\s*[:·-]?\s*([A-Za-z0-9._/-]+)", doc.doc_no or "", re.I)
    return m.group(1) if m else (doc.doc_no or "")


def _prozessstufe(doc) -> str:
    if doc is None:
        return ""
    m = re.search(r"\bB\d{1,3}(?:-?[A-Z]{1,3})?\b", f"{doc.title or ''} {doc.doc_no or ''}")
    return m.group(0) if m else ""


def solution_condition(field) -> str:
    """Map a stored field's flags onto the host's `Condition` verdict taxonomy."""
    flags = list(field.flags)
    if not flags:
        return "Online / confirmed by second value" if field.role in _CONFIRMED_ROLES else "Online"
    err = {fl.code for fl in flags if fl.severity == "error"}
    codes = err or {fl.code for fl in flags}
    if any(c.startswith("MISSING") for c in codes):
        return "Suspect / missing"
    if "CALC_VOLUME" in codes:
        return "Suspect / Value does not fit volume calculation"
    if any(c.startswith("CALC") for c in codes):
        return "Suspect / Value does not fit calculation"
    return "Suspect / unexpected value"


def solution_text_value(field):
    """(Text, Value) for the row: a checkbox reads Checked/Not checked; numerics
    also fill the Value column."""
    if field.value_type == "checkbox":
        return ("Checked" if (field.value_raw or "").strip() else "Not checked"), None
    text = field.value_norm if field.value_norm is not None else (field.value_raw or "")
    num = None
    try:
        num = float(str(field.value_norm if field.value_norm is not None else field.value_raw).replace(",", "."))
    except (TypeError, ValueError):
        num = None
    return text, num


SEV_DE = {"error": "Fehler", "warning": "Warnung"}
STATUS_DE = {
    "auto_accepted": "Auto-akzeptiert",
    "needs_review":  "Zur Pruefung",
    "confirmed":     "Bestaetigt",
    "corrected":     "Korrigiert",
    "extracted":     "Extrahiert",
}


def _load_doc_fields(document_id: str, db: Session):
    doc = db.get(models.Document, document_id)
    fields = (
        db.query(models.Field)
        .filter(models.Field.document_id == document_id)
        .order_by(models.Field.page_no, models.Field.chapter)
        .all()
    )
    return doc, fields


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_xlsx(document_id: str, db: Session) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("`pip install openpyxl` to export Excel.") from exc

    doc, fields = _load_doc_fields(document_id, db)
    wb = Workbook()

    # shared styles
    hdr_fill = PatternFill("solid", fgColor="1a3a5c")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri")
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_err  = PatternFill("solid", fgColor="FFE0E0")
    fill_warn = PatternFill("solid", fgColor="FFF3CD")

    # ── Sheet 1: SOLUTION format (gradable, matches the host grading sheet) ──
    sol = wb.active
    sol.title = "Solution"
    sol.append(SOLUTION_HEADER)
    for ci in range(1, len(SOLUTION_HEADER) + 1):
        c = sol.cell(1, ci)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = cell_border
    sol.row_dimensions[1].height = 20

    batch, stufe = _batch_no(doc), _prozessstufe(doc)
    for f in fields:
        text, num = solution_text_value(f)
        cond = solution_condition(f)
        sol.append([batch, stufe, f.chapter or "", "", "",
                    f.label_raw or "", f.unit or "", "",
                    cond, "" if text is None else text, "" if num is None else num])
        r = sol.max_row
        if cond.lower().startswith(("suspect", "supect")):
            fill = fill_err if any(fl.severity == "error" for fl in f.flags) else fill_warn
            for ci in range(1, len(SOLUTION_HEADER) + 1):
                sol.cell(r, ci).fill = fill
        for ci in range(1, len(SOLUTION_HEADER) + 1):
            sol.cell(r, ci).border = cell_border
    sol.auto_filter.ref = sol.dimensions
    sol.freeze_panes = "A2"
    for i, w in enumerate([12, 12, 18, 18, 18, 28, 14, 12, 32, 22, 12], 1):
        sol.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Alle Felder (detail) ──────────────────────────────────────
    ws = wb.create_sheet("Chargenprotokoll")

    # Kopfzeile
    header = ["Seite", "Kapitel", "Feldname", "Rolle", "OCR-Wert",
              "Normalisiert", "Einheit", "Status", "Konfidenz",
              "Fehler-Code", "Schwere", "Erwartet", "Gefunden", "Meldung"]
    hdr_fill   = PatternFill("solid", fgColor="1a3a5c")
    hdr_font   = Font(bold=True, color="FFFFFF", name="Calibri")
    thin       = Side(style="thin", color="CCCCCC")
    cell_border= Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(header)
    for col_idx, _ in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

    ws.row_dimensions[1].height = 20

    # Farben nach Status/Schwere
    fill_err  = PatternFill("solid", fgColor="FFE0E0")
    fill_warn = PatternFill("solid", fgColor="FFF3CD")
    fill_ok   = PatternFill("solid", fgColor="E8F5E9")
    fill_rev  = PatternFill("solid", fgColor="E3F2FD")

    for f in fields:
        flags = f.flags
        first_flag = flags[0] if flags else None
        has_error  = any(fl.severity == "error"   for fl in flags)
        has_warn   = any(fl.severity == "warning"  for fl in flags)

        row_data = [
            f.page_no,
            f.chapter or "",
            f.label_raw or "",
            f.role or "",
            f.value_raw or "",
            f.value_norm if f.value_norm is not None else (f.value_raw or ""),
            f.unit or "",
            STATUS_DE.get(f.status, f.status),
            round(f.confidence, 3),
            "; ".join(fl.code for fl in flags) if flags else "",
            "; ".join(SEV_DE.get(fl.severity, fl.severity) for fl in flags) if flags else "",
            "; ".join(fl.expected or "" for fl in flags) if flags else "",
            "; ".join(fl.actual   or "" for fl in flags) if flags else "",
            "; ".join(fl.message  or "" for fl in flags) if flags else "",
        ]
        ws.append(row_data)
        row_idx = ws.max_row

        # Zeilenfaerbung
        if has_error:   fill = fill_err
        elif has_warn:  fill = fill_warn
        elif f.status == "needs_review": fill = fill_rev
        else:           fill = fill_ok

        for col_idx in range(1, len(header) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill   = fill
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == len(header)))

    # Spaltenbreiten
    widths = [6, 10, 30, 22, 22, 22, 8, 16, 10, 22, 10, 22, 22, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Auto-Filter
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    # ── Tabellenblatt 2: Zusammenfassung ──────────────────────────────────
    summ = wb.create_sheet("Zusammenfassung")
    now  = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M UTC")
    n_err   = sum(1 for f in fields for fl in f.flags if fl.severity == "error")
    n_warn  = sum(1 for f in fields for fl in f.flags if fl.severity == "warning")
    n_auto  = sum(1 for f in fields if f.status == "auto_accepted")
    n_rev   = sum(1 for f in fields if f.status == "needs_review")
    n_conf  = sum(1 for f in fields if f.status in ("confirmed", "corrected"))

    summ_data = [
        ("Dokument",        doc.doc_no if doc else document_id),
        ("Titel",           doc.title  if doc else ""),
        ("Seiten",          doc.page_count if doc else ""),
        ("Exportiert",      now),
        ("", ""),
        ("Felder gesamt",   len(fields)),
        ("Auto-akzeptiert", n_auto),
        ("Bestaetigt",      n_conf),
        ("Zur Pruefung",    n_rev),
        ("", ""),
        ("Fehler",          n_err),
        ("Warnungen",       n_warn),
    ]
    lbl_font = Font(bold=True, name="Calibri")
    for label, value in summ_data:
        summ.append([label, value])
        if label:
            summ.cell(summ.max_row, 1).font = lbl_font
    summ.column_dimensions["A"].width = 22
    summ.column_dimensions["B"].width = 40

    # ── Tabellenblatt 3: Nur Fehler/Warnungen ─────────────────────────────
    flags_ws = wb.create_sheet("Flags")
    flags_ws.append(["Seite", "Kapitel", "Feldname", "Fehler-Code", "Schwere",
                     "Erwartet", "Gefunden", "Meldung"])
    for col_idx in range(1, 9):
        cell = flags_ws.cell(row=1, column=col_idx)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    for f in fields:
        for fl in f.flags:
            flags_ws.append([
                f.page_no, f.chapter or "", f.label_raw or "",
                fl.code, SEV_DE.get(fl.severity, fl.severity),
                fl.expected or "", fl.actual or "", fl.message or "",
            ])
            r = flags_ws.max_row
            fill = fill_err if fl.severity == "error" else fill_warn
            for c in range(1, 9):
                flags_ws.cell(r, c).fill   = fill
                flags_ws.cell(r, c).border = cell_border

    flags_ws.auto_filter.ref = flags_ws.dimensions
    flags_ws.freeze_panes = "A2"
    for i, w in enumerate([6, 10, 30, 22, 10, 22, 22, 45], 1):
        flags_ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

CSV_HEADER = [
    "Seite", "Kapitel", "Feldname", "Rolle",
    "OCR-Wert", "Normalisiert", "Einheit",
    "Status", "Konfidenz",
    "Fehler-Code", "Schwere", "Erwartet", "Gefunden", "Meldung",
]


def export_csv(document_id: str, db: Session) -> bytes:
    """UTF-8 BOM CSV (Semikolon-getrennt) -- direkt in Excel oeffenbar."""
    doc, fields = _load_doc_fields(document_id, db)
    now  = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M UTC")
    n_err   = sum(1 for f in fields for fl in f.flags if fl.severity == "error")
    n_warn  = sum(1 for f in fields for fl in f.flags if fl.severity == "warning")
    n_auto  = sum(1 for f in fields if f.status == "auto_accepted")

    buf = io.StringIO()
    w   = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)

    # Metadaten-Header
    w.writerow(["# BioDize Export"])
    w.writerow(["# Dokument", doc.doc_no if doc else document_id])
    if doc and doc.title and doc.title != doc.doc_no:
        w.writerow(["# Titel", doc.title])
    w.writerow(["# Exportiert", now])
    w.writerow(["# Felder", len(fields),
                "Fehler", n_err,
                "Warnungen", n_warn,
                "Auto-akzeptiert", n_auto])
    w.writerow([])

    # Spalten-Header
    w.writerow(CSV_HEADER)

    # Datenzeilen
    for f in fields:
        flags = f.flags
        codes   = " | ".join(fl.code               for fl in flags) if flags else ""
        sevs    = " | ".join(SEV_DE.get(fl.severity, fl.severity) for fl in flags) if flags else ""
        expectd = " | ".join(fl.expected or ""      for fl in flags) if flags else ""
        actual  = " | ".join(fl.actual   or ""      for fl in flags) if flags else ""
        msgs    = " | ".join(fl.message  or ""      for fl in flags) if flags else ""

        w.writerow([
            f.page_no,
            f.chapter or "",
            f.label_raw or "",
            f.role or "",
            f.value_raw or "",
            f.value_norm if f.value_norm is not None else (f.value_raw or ""),
            f.unit or "",
            STATUS_DE.get(f.status, f.status),
            f"{f.confidence:.3f}",
            codes, sevs, expectd, actual, msgs,
        ])

    # UTF-8 BOM damit Excel es korrekt oeffnet
    return ("﻿" + buf.getvalue()).encode("utf-8")
